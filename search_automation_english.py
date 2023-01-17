from googlesearch import search
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Cargar la base a Python y seleccionar la hoja
loc = (r'C:\Users\diego\Documents\SV\Python Search Automation\Alcaldes python.xlsx')
wb = load_workbook(loc)
df = wb['Hoja3']

# Seleccionar la columna de nombres como una variable
nombres = df['A']

for cell in nombres:
    # Imprimir el nombre de cada candidato y guardarlo en la variable candidato
    candidato = (cell.value)
    print(candidato)

d = []
for cell in nombres:
    # Para cada celda del excel, insertarla en el query de búsqueda
    query = [cell.value + ' Alcalde'' represión', cell.value + ' Alcalde'' detenido', cell.value +' Alcalde'' arresto', cell.value + ' Alcalde'' inhabilitado']
    print(query)
    for i in query:
        # Para cada query de búsqueda, retorna los links de la búsqueda de google
        busqueda = search(i, tld='com', lang='es', tbs='0', safe='off', num=2, start=0, stop=2, pause=2.0, country='', extra_params=None, user_agent=None)
        for j in busqueda:
            d.append(
                {
                    'Links':j
                }
            )

pd.DataFrame(d)
print(d)

d.to_csv()